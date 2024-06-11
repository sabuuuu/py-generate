from pulp import *
import openpyxl
import pandas as pd

class EmploiDuTemps:
    def __init__(self, K, M, L, H, S ,days_of_week):
        self.K = K  # Nombre de groupes
        self.M = M  # Nombre d'enseignements
        self.L = L  # Nombre d'enseignants
        self.H = H  # Nombre de créneaux horaires par jour
        self.S = S  # Nombre de salles
        self.days_of_week = days_of_week  # Nombre de jours de la semaine
        # Initialisation du problème
        self.prob = LpProblem("Emploi_du_temps", LpMinimize)

        # Variables de décision
        self.x = LpVariable.dicts("x", [(k, h, l, m, s) for k in range(1, K+1) 
                                        for h in range(1, H+1) 
                                        for l in range(1, L+1) 
                                        for m in range(1, M+1) 
                                        for s in range(1, S+1)], 
                                  cat='Binary')

    def ajouter_contraintes(self):
        # Contrainte 1: Chaque enseignement doit être suivi par chaque groupe
        for k in range(1, self.K+1):
            for m in range(1, self.M+1):
                self.prob += lpSum([self.x[(k, h, l, m, s)] 
                                    for h in range(1, self.H+1) 
                                    for l in range(1, self.L+1) 
                                    for s in range(1, self.S+1)]) == 1

        # Contrainte 2: Un enseignant ne peut enseigner qu'un seul enseignement à la fois
        for l in range(1, self.L+1):
            for h in range(1, self.H+1):
                self.prob += lpSum([self.x[(k, h, l, m, s)] 
                                    for k in range(1, self.K+1) 
                                    for m in range(1, self.M+1)
                                    for s in range (1,self.S+1)]) <= 1

        # Contrainte 3: Un groupe ne peut suivre qu'un seul enseignement à la fois
        for k in range(1, self.K+1):
            for h in range(1, self.H+1):
                self.prob += lpSum([self.x[(k, h, l, m, s)] 
                                    for l in range(1, self.L+1) 
                                    for m in range(1, self.M+1)
                                    for s in range(1, self.S+1)]) <= 1

        # Contrainte 4: Chaque salle ne peut être occupée que par un seul groupe à chaque créneau horaire
        for h in range(1, self.H+1):
            for s in range(1, self.S+1):
                self.prob += lpSum([self.x[(k, h, l, m, s)] 
                                    for k in range(1, self.K+1) 
                                    for l in range(1, self.L+1) 
                                    for m in range(1, self.M+1)]) <= 1

    def minimiser_utilisation_dernier_creneau(self):
        # Objectif: Minimiser l'utilisation du dernier créneau du jour
        self.prob += lpSum([self.x[(k, h, l, m, s)] 
                            for k in range(1, self.K+1) 
                            for l in range(1, self.L+1) 
                            for m in range(1, self.M+1)
                            for s in range(1, self.S+1)
                            for h in  [6, 12, 18, 24, 30]])  

    def resoudre(self):
        self.ajouter_contraintes()
        self.minimiser_utilisation_dernier_creneau()
        # Résolution du problème
        self.prob.solve()
        status = self.prob.status
        print('status = ', status)
        if(status == 1):
            # Affichage de la solution
            print("Utilisation des derniers créneaux du jour = ", value(self.prob.objective))
            for v in self.prob.variables():
                if(v.varValue == 1):
                    print(v.name, "=", v.varValue)
        elif(status == -1):
            print('Le probleme est infaisable.')
        
        return self.prob
    def creer_emploi_du_temps_excel(self, course_names, teacher_names, group_names, room_names, time_slots):
      wb = openpyxl.Workbook()
      sheet = wb.active
      # Set initial column widths (adjust as needed)
      for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

      # Set initial row heights (adjust as needed)
      sheet.row_dimensions[1].height = 30  # Header row with time slots

      # Create headers for days of the week (assuming days_of_week is a list)
      for col_index, day in enumerate(self.days_of_week, start=2):
        sheet.cell(row=1, column=col_index).value = day
        sheet.cell(row=1, column=col_index).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

      # Create headers for time slots (assuming time_slots is a list)
      for row_index, time_slot in enumerate(time_slots, start=2):
        sheet.cell(row=row_index, column=1).value = time_slot
        sheet.cell(row=row_index, column=1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

      # Iterate through time slots, rooms to find assigned courses
      for row_index, time_slot in enumerate(time_slots, start=2):
        for col_index, room_index in enumerate(range(1, self.S + 1), start=2):
          # Find assigned course for this time slot and room
          assigned_course = None
          for k in range(1, self.K + 1):
            for l in range(1, self.L + 1):
              for m in range(1, self.M + 1):
                if self.x[(k, time_slot, l, m, room_index)].value() == 1.0:
                  assigned_course = (k, l, m)
                  break  # Exit inner loops once a course is found

          # Fill in the timetable cell if a course is assigned
          if assigned_course:
            group_index, teacher_index, course_index = assigned_course
            course_info = f"{course_names[course_index - 1]} ({teacher_names[teacher_index - 1]}) - {group_names[group_index - 1]} - {room_names[room_index - 1]}"
            sheet.cell(row=row_index, column=col_index).value = course_info
            # Adjust column widths for content (optional)
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = max(sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width, len(course_info) + 5)

          # Set column widths
          for col in range(1, sheet.max_column+1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

          # Set row heights
          for row in range(1, sheet.max_row+1):
            sheet.row_dimensions[row].height = 30

          # Set font size
          for row in range(1, sheet.max_row+1):
            for col in range(1, sheet.max_column+1):
              sheet.cell(row, col).font = openpyxl.styles.Font(size=12)

          # Set borders
          for row in range(1, sheet.max_row+1):
            for col in range(1, sheet.max_column+1):
              sheet.cell(row, col).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                                  right=openpyxl.styles.Side(style='thin'),
                                                                  top=openpyxl.styles.Side(style='thin'),
                                                                  bottom=openpyxl.styles.Side(style='thin'))

          # Set alignment
          for row in range(1, sheet.max_row+1):
            for col in range(1, sheet.max_column+1):
              sheet.cell(row, col).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

          # Set background color
          for row in range(1, sheet.max_row+1):
            for col in range(1, sheet.max_column+1):
              sheet.cell(row, col).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='D3D3D3')
              
          
          # Save the Excel file (replace 'emploi_du_temps.xlsx' with your desired filename)
          wb.save("emploi_du_temps.xlsx")

      

K = 3  # Number of groups
M = 6  # Number of courses
L = 6  # Number of teachers (increased to match teacher_names)
H = 30  # Number of time slots per week
S = 3  # Number of rooms

days_of_week = ["Dimanche", "Lundi", "Mardi", "Mercredi", "Jeudi"]
course_names = ["Pl", "Pa","PLNE","TG","Anglais","Math"]
teacher_names = ["Teacher 1", "Teacher 2", "Teacher 3", "Teacher 4", "Teacher 5", "Teacher 6"]
group_names = ["Groupe 1", "Groupe 2","Groupe 3"]
rooms = ["Salle 1", "Salle 2", "Salle 3"]
time_slots = range(1, H+1)

# emploi_du_temps = EmploiDuTemps(K, M, L, H, S)
# prob = emploi_du_temps.resoudre()


edt = EmploiDuTemps(K, M, L, H, S, days_of_week)
edt.ajouter_contraintes()
edt.resoudre()
edt.creer_emploi_du_temps_excel(course_names, teacher_names, group_names, rooms, time_slots)
